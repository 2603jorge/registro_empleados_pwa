import os
import base64
import json
import requests
from datetime import datetime
from urllib.parse import urlparse
from flask import Flask, request, render_template, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
os.makedirs("static/fotos", exist_ok=True)

# ===== Variables de entorno =====
CLIENT_ID = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]
TENANT_ID = os.environ["TENANT_ID"]
# p.ej. https://agricactus2.sharepoint.com/sites/CALIDAD
SHAREPOINT_SITE = os.environ["SHAREPOINT_SITE"]
# p.ej. "DOCUMENTOS"
SHAREPOINT_LIBRARY = os.environ.get("SHAREPOINT_LIBRARY", "DOCUMENTOS").strip()
# p.ej. subcarpeta dentro de la biblioteca, o vacío para raíz
SHAREPOINT_FOLDER = os.environ.get("SHAREPOINT_FOLDER", "").strip()
# p.ej. "registro_empleados.xlsx"
SHAREPOINT_DOC = os.environ["SHAREPOINT_DOC"]

# ===== Ctes =====
GRAPH = "https://graph.microsoft.com/v1.0"

# ===== Token cache sencillo =====
_site_id_cache = {"value": None}

def obtener_token():
    """Obtiene un access token (client_credentials) para Microsoft Graph."""
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {
        "client_id": CLIENT_ID,
        "scope": "https://graph.microsoft.com/.default",
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
    }
    r = requests.post(url, headers=headers, data=data, timeout=30)
    if r.status_code != 200:
        print("Error al obtener token:", r.text)
        raise Exception("No se pudo obtener access_token")
    return r.json()["access_token"]

def normalizar_site_url(site_url: str):
    """
    Devuelve hostname y el path del SITIO solamente, p.ej.:
    https://agricactus2.sharepoint.com/sites/CALIDAD/SitePages/CollabHome.aspx
    -> ('agricactus2.sharepoint.com', '/sites/CALIDAD')
    """
    p = urlparse(site_url.strip())
    hostname = p.netloc
    partes = [seg for seg in p.path.split("/") if seg]
    site_path = ""
    if len(partes) >= 2 and partes[0].lower() == "sites":
        site_path = f"/sites/{partes[1]}"
    else:
        site_path = p.path if p.path.startswith("/") else f"/{p.path}"
    return hostname, site_path

def obtener_site_id():
    """Resuelve el siteId usando /v1.0/sites/{hostname}:/{site-path}?$select=id."""
    if _site_id_cache["value"]:
        return _site_id_cache["value"]
    token = obtener_token()
    hostname, site_path = normalizar_site_url(SHAREPOINT_SITE)
    url = f"{GRAPH}/sites/{hostname}:/{site_path}?$select=id,webUrl"
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code != 200:
        print("Error al obtener siteId:", r.text)
        raise Exception("No se pudo resolver el siteId de SharePoint")
    site_id = r.json().get("id")
    if not site_id:
        raise Exception("Respuesta de Graph no contiene 'id' para el sitio")
    _site_id_cache["value"] = site_id
    return site_id

def obtener_drive_id(site_id: str, library_name: str):
    """
    Busca el drive (biblioteca) por displayName/name (p.ej. 'DOCUMENTOS').
    """
    token = obtener_token()
    url = f"{GRAPH}/sites/{site_id}/drives"
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code != 200:
        print("Error listando drives:", r.text)
        raise Exception("No se pudieron obtener las bibliotecas del sitio")

    data = r.json().get("value", [])
    for d in data:
        if d.get("displayName") == library_name or d.get("name") == library_name:
            return d["id"]

    nombres = [f"{d.get('displayName')}|{d.get('name')}" for d in data]
    raise Exception(f"No se encontró la biblioteca '{library_name}'. Disponibles: {nombres}")

def subir_a_sharepoint(ruta_excel_local: str):
    """
    Sube/reemplaza el archivo en la biblioteca indicada (SHAREPOINT_LIBRARY).
    Usa: PUT /drives/{driveId}/root:/carpeta/archivo:/content
    """
    token = obtener_token()
    site_id = obtener_site_id()
    drive_id = obtener_drive_id(site_id, SHAREPOINT_LIBRARY)

    # Construimos la ruta destino dentro de la biblioteca
    if SHAREPOINT_FOLDER:
        drive_path = f"/{SHAREPOINT_FOLDER.strip('/')}/{SHAREPOINT_DOC}"
    else:
        drive_path = f"/{SHAREPOINT_DOC}"

    url = f"{GRAPH}/drives/{drive_id}/root:{drive_path}:/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    with open(ruta_excel_local, "rb") as f:
        r = requests.put(url, headers=headers, data=f.read(), timeout=60)

    if r.status_code not in (200, 201):
        print("Error al subir a SharePoint:", r.status_code, r.text)
        return False
    return True

def guardar_archivo(nombre_base: str, base64_data: str):
    """
    Guarda un archivo enviado en base64 (imagen/pdf) en static/fotos y devuelve el nombre.
    Si no hay datos, devuelve cadena vacía.
    """
    if not base64_data:
        return ""
    try:
        ext = ".jpg"
        if ";base64," in base64_data:
            meta, b64 = base64_data.split(",", 1)
            if "pdf" in meta:
                ext = ".pdf"
        else:
            b64 = base64_data

        nombre_archivo = f"{nombre_base}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
        ruta = os.path.join("static", "fotos", nombre_archivo)
        with open(ruta, "wb") as f:
            f.write(base64.b64decode(b64))
        return nombre_archivo
    except Exception as e:
        print("Error guardando archivo base64:", e)
        return ""

def asegurar_excel_local(path_excel: str):
    """Si no existe el Excel local, lo crea con encabezados."""
    if os.path.exists(path_excel):
        return
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Fecha", "Nombre", "Edad", "CURP", "RFC", "NSS", "Teléfono",
        "Dirección", "Leer/Escribir", "Discapacidad", "Experiencia",
        "Salud", "Origen", "Observaciones", "Trabajo previo",
        "Año trabajo", "Área trabajo", "Contacto emergencia",
        "Teléfono emergencia", "INE Frente", "INE Reverso", "CURP Archivo",
        "Acta/Comprobante", "Foto facial"
    ])
    wb.save(path_excel)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            datos = request.get_json(silent=True) or request.form

            # Si llegan archivos como <input type="file"> y no en base64:
            archivos = request.files if request.files else None

            archivos_nombres = {
                "ine_frente": guardar_archivo("ine_frente", datos.get("ine_frente")) if not archivos else "",
                "ine_reverso": guardar_archivo("ine_reverso", datos.get("ine_reverso")) if not archivos else "",
                "curp_archivo": guardar_archivo("curp_archivo", datos.get("curp_archivo")) if not archivos else "",
                "documentos": guardar_archivo("documentos", datos.get("documentos_base64")) if not archivos else "",
                "foto": guardar_archivo("foto", datos.get("foto_base64")) if not archivos else "",
            }

            # Si vinieron como archivos tradicionales, guárdalos también en static/fotos
            if archivos:
                for key in ["ine_frente", "ine_reverso", "curp_archivo", "documentos", "foto"]:
                    if key in archivos and archivos[key].filename:
                        filename = f"{key}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{archivos[key].filename}"
                        save_path = os.path.join("static", "fotos", filename)
                        archivos[key].save(save_path)
                        archivos_nombres[key] = filename

            fila = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                datos.get("nombre", ""), datos.get("edad", ""), datos.get("curp", ""),
                datos.get("rfc", ""), datos.get("nss", ""), datos.get("telefono", ""),
                datos.get("direccion", ""), datos.get("leer_escribir", ""),
                datos.get("discapacidad", ""), datos.get("experiencia", ""),
                datos.get("salud", ""), datos.get("origen", ""),
                datos.get("observaciones", ""), datos.get("trabajo_previo", ""),
                datos.get("año_trabajo", ""), datos.get("area_trabajo", ""),
                datos.get("contacto_emergencia", ""), datos.get("telefono_emergencia", ""),
                archivos_nombres["ine_frente"], archivos_nombres["ine_reverso"],
                archivos_nombres["curp_archivo"], archivos_nombres["documentos"],
                archivos_nombres["foto"]
            ]

            archivo_excel_local = "registro_empleados.xlsx"
            asegurar_excel_local(archivo_excel_local)

            wb = load_workbook(archivo_excel_local)
            ws = wb.active
            ws.append(fila)
            wb.save(archivo_excel_local)

            if subir_a_sharepoint(archivo_excel_local):
                return jsonify({"ok": True}), 200
            else:
                return jsonify({"error": "No se pudo subir a SharePoint"}), 500

        except Exception as e:
            print("Error en POST /:", e)
            return jsonify({"error": str(e)}), 500

    return render_template("index.html")

@app.route("/service-worker.js")
def sw():
    return send_from_directory("static", "service-worker.js")

@app.route("/manifest.json")
def manifest():
    return send_from_directory("static", "manifest.json")

if __name__ == "__main__":
    # En Render se usa gunicorn; esto es para pruebas locales.
    app.run(host="0.0.0.0", port=5000)

